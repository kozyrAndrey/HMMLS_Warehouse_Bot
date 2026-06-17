from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")


def create_applications_xlsx(applications, output_path):
    output_path = Path(output_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Резюме"

    headers = [
        "ID заявки",
        "Дата",
        "Telegram ID",
        "Username",
        "Telegram имя",
        "Должность",
        "ФИО",
        "Возраст",
        "Проживание",
        "Опыт",
        "Смен в неделю",
        "Часов за смену",
        "Статус",
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, application in enumerate(applications, start=2):
        created_at = application.get("created_at")
        created_at_text = created_at.strftime("%d.%m.%Y %H:%M") if created_at else ""
        username = application.get("telegram_username") or ""
        username_text = f"@{username}" if username else ""

        values = [
            application.get("id"),
            created_at_text,
            application.get("telegram_user_id"),
            username_text,
            application.get("telegram_full_name"),
            application.get("position"),
            application.get("full_name"),
            application.get("age"),
            application.get("settlement"),
            application.get("experience"),
            application.get("shifts_per_week"),
            application.get("hours_per_shift"),
            application.get("status"),
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=col == 10)

    widths = [12, 18, 16, 22, 28, 18, 32, 10, 26, 60, 16, 18, 14]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    return output_path
