import re
import warnings
from copy import copy
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from config import (
    MARKING_ONE_C_CONSIGNOR,
    MARKING_ONE_C_DEFAULT_GENDER,
    MARKING_ONE_C_RETAIL_PRICE_TYPE,
    MARKING_ONE_C_TEMPLATE_PATH,
)
from modules.marking.one_c_config import (
    ALLOWED_CATEGORIES,
    ALLOWED_GENDERS,
    CATEGORY_KEYWORDS,
    COUNTRY_ALIASES,
    FIELD_ALIASES,
    GENDER_ALIASES,
)
from modules.marking.storage import normalize_gtin


ONE_C_SHEET_NAME = "Список товаров"
ONE_C_HEADERS = [
    "№",
    "Артикул",
    "Наименование",
    "Пол",
    "Страна происхождения",
    "Бренд",
    "Производитель",
    "Состав",
    "Размер",
    "Цвет",
    "Розничная цена (руб)",
    "Кол-во (справочно)",
    "доп. поле (штрих-код)",
    "Категория",
]
BRAND_NAME = "Hommeplusless"
FIXED_CATEGORY = "Товары легкой промышленности"


class OneCExportValidationError(RuntimeError):
    def __init__(self, errors):
        self.errors = list(errors)
        super().__init__("Не удалось сформировать Excel для 1С.")


@dataclass
class OneCExportItem:
    article: str
    honest_sign_name: str
    gender: str
    country_of_origin: str
    brand: str
    manufacturer: str
    composition: str
    size: str
    color: str
    retail_price: Decimal
    marking_code_count: int
    ean13: str
    category: str
    gtin: str


def normalize_lookup_text(value):
    normalized = str(value or "").strip().casefold().replace("ё", "е")
    normalized = re.sub(r"[.,;:()\[\]{}_/\\-]+", " ", normalized)
    return " ".join(normalized.split())


def named_value(items, aliases):
    aliases = {normalize_lookup_text(alias) for alias in aliases}
    for item in items or []:
        if not isinstance(item, dict):
            continue
        if normalize_lookup_text(item.get("name")) not in aliases:
            continue
        value = item.get("value")
        if isinstance(value, dict):
            value = value.get("name") or value.get("value")
        value = str(value or "").strip()
        if value:
            return value
    return ""


def field_value(assortment, product, field_name, fallback_alias_group=None):
    aliases = FIELD_ALIASES[field_name]
    for source in (
        assortment.get("characteristics"),
        assortment.get("attributes"),
        product.get("characteristics"),
        product.get("attributes"),
    ):
        value = named_value(source, aliases)
        if value:
            return value
    if fallback_alias_group:
        return field_value(assortment, product, fallback_alias_group)
    return ""


def normalize_gender(value, default_gender=MARKING_ONE_C_DEFAULT_GENDER):
    source = str(value or default_gender or "").strip()
    normalized = normalize_lookup_text(source)
    result = GENDER_ALIASES.get(normalized)
    if result:
        return result
    for allowed in ALLOWED_GENDERS:
        if normalize_lookup_text(allowed) == normalized:
            return allowed
    raise ValueError(f"не удалось сопоставить значение пола «{source or '<пусто>'}»")


def normalize_category(value):
    source = str(value or "").strip()
    normalized = normalize_lookup_text(source)
    for allowed in ALLOWED_CATEGORIES:
        if normalize_lookup_text(allowed) == normalized:
            return allowed

    matches = []
    for category, keywords in CATEGORY_KEYWORDS.items():
        if any(normalize_lookup_text(keyword) in normalized for keyword in keywords):
            matches.append(category)
    if len(set(matches)) == 1:
        return matches[0]
    raise ValueError(f"не удалось сопоставить категорию «{source or '<пусто>'}»")


def normalize_country(value):
    source = str(value or "").strip()
    if not source:
        return ""
    return COUNTRY_ALIASES.get(normalize_lookup_text(source), source.upper())


def extract_country(assortment, product):
    value = field_value(assortment, product, "country")
    if value:
        return normalize_country(value)
    country = product.get("country") or assortment.get("country") or {}
    if isinstance(country, dict):
        return normalize_country(country.get("name"))
    return normalize_country(country)


def extract_ean13(assortment, product):
    values = []
    for source in (assortment.get("barcodes"), product.get("barcodes")):
        for barcode in source or []:
            if not isinstance(barcode, dict):
                continue
            value = str(barcode.get("ean13") or "").strip()
            if value and value not in values:
                values.append(value)
        if values:
            break
    if not values:
        raise ValueError("отсутствует EAN-13")
    if len(values) > 1:
        raise ValueError("найдено несколько разных EAN-13: " + ", ".join(values))
    if not re.fullmatch(r"\d{13}", values[0]):
        raise ValueError(f"EAN-13 «{values[0]}» должен содержать ровно 13 цифр")
    return values[0]


def price_from_sale_prices(sale_prices, wanted_name=None, contains_retail=False):
    wanted = normalize_lookup_text(wanted_name)
    for sale_price in sale_prices or []:
        if not isinstance(sale_price, dict):
            continue
        actual = normalize_lookup_text((sale_price.get("priceType") or {}).get("name"))
        matches = actual == wanted if wanted else (contains_retail and "рознич" in actual)
        if not matches:
            continue
        value = sale_price.get("value")
        try:
            result = Decimal(str(value)) / Decimal("100")
        except (InvalidOperation, TypeError, ValueError):
            raise ValueError(f"розничная цена «{value}» имеет некорректный формат")
        if not result.is_finite() or result < 0:
            raise ValueError(f"розничная цена «{value}» имеет некорректный формат")
        return result
    return None


def extract_retail_price(assortment, product, configured_type=MARKING_ONE_C_RETAIL_PRICE_TYPE):
    sources = (assortment.get("salePrices"), product.get("salePrices"))
    for source in sources:
        value = price_from_sale_prices(source, wanted_name="Розничная цена")
        if value is not None:
            return value
    for source in sources:
        value = price_from_sale_prices(source, contains_retail=True)
        if value is not None:
            return value
    for source in sources:
        value = price_from_sale_prices(source, wanted_name=configured_type)
        if value is not None:
            return value
    raise ValueError(f"отсутствует тип цены «{configured_type}»")


def build_error_prefix(article, name, gtin, size):
    parts = []
    if article:
        parts.append(f"артикул {article}")
    if name:
        parts.append(name)
    if gtin:
        parts.append(f"GTIN {gtin}")
    if size:
        parts.append(f"размер {size}")
    return ", ".join(parts) or "Неизвестная позиция"


def build_one_c_export_items(rows, catalog_names):
    catalog = {}
    for gtin, name in (catalog_names or {}).items():
        try:
            catalog[normalize_gtin(gtin)] = str(name or "").strip()
        except ValueError:
            continue

    errors = []
    grouped = {}
    gtin_owners = {}
    code_owners = {}

    for position_number, row in enumerate(rows, start=1):
        codes = {
            str(code)
            for code in row.get("codes") or []
            if code is not None and str(code)
        }
        if not codes:
            continue

        assortment = row.get("assortment") or {}
        product = row.get("product") or {}
        article = str(row.get("article") or "").strip()
        raw_gtin = str(row.get("gtin") or "").strip()
        normalized_gtin = ""
        try:
            normalized_gtin = normalize_gtin(raw_gtin)
        except ValueError:
            pass
        honest_sign_name = catalog.get(normalized_gtin, "")
        size = field_value(assortment, product, "size")
        color = field_value(assortment, product, "color")
        prefix = build_error_prefix(article, honest_sign_name or str(row.get("name") or ""), raw_gtin, size)
        position_errors = []

        if not article:
            position_errors.append("отсутствует артикул")
        elif len(article) > 50:
            position_errors.append("длина артикула превышает 50 символов")
        if not raw_gtin:
            position_errors.append("отсутствует GTIN")
        elif not normalized_gtin:
            position_errors.append(f"GTIN «{raw_gtin}» имеет некорректный формат")
        if normalized_gtin and not honest_sign_name:
            position_errors.append("GTIN отсутствует в локальном справочнике Честного ЗНАКа")

        raw_gender = field_value(assortment, product, "gender")
        try:
            gender = normalize_gender(raw_gender)
        except ValueError as error:
            position_errors.append(str(error))
            gender = ""

        country = extract_country(assortment, product)
        manufacturer = field_value(assortment, product, "manufacturer")
        composition = field_value(assortment, product, "composition") or field_value(
            assortment, product, "material"
        )
        category = FIXED_CATEGORY

        try:
            ean13 = extract_ean13(assortment, product)
        except ValueError as error:
            position_errors.append(str(error))
            ean13 = ""

        try:
            retail_price = extract_retail_price(assortment, product)
        except ValueError as error:
            position_errors.append(str(error))
            retail_price = None

        for field_name, value in (
            ("страна происхождения", country),
            ("производитель", manufacturer),
            ("состав", composition),
            ("размер", size),
            ("цвет", color),
        ):
            if not value:
                position_errors.append(f"отсутствует {field_name}")

        if position_errors:
            errors.extend(f"{prefix}: {message}." for message in position_errors)
            continue

        retail_price = retail_price.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        key = (article, normalized_gtin, size, color, ean13)
        previous_key = gtin_owners.get(normalized_gtin)
        if previous_key is not None and previous_key != key:
            errors.append(f"{prefix}: одному GTIN соответствуют разные товарные модификации.")
            continue
        gtin_owners[normalized_gtin] = key

        conflicting_code = False
        for code in codes:
            owner = code_owners.get(code)
            if owner is not None and owner != key:
                errors.append(f"{prefix}: КИЗ {code} относится к нескольким товарным модификациям.")
                conflicting_code = True
            else:
                code_owners[code] = key
        if conflicting_code:
            continue

        if key not in grouped:
            grouped[key] = {
                "article": article,
                "honest_sign_name": honest_sign_name,
                "gender": gender,
                "country_of_origin": country,
                "brand": BRAND_NAME,
                "manufacturer": manufacturer,
                "composition": composition,
                "size": size,
                "color": color,
                "retail_price": retail_price,
                "ean13": ean13,
                "category": category,
                "gtin": normalized_gtin,
                "codes": set(),
                "position_number": position_number,
            }
        grouped[key]["codes"].update(codes)

    if not grouped and not errors:
        errors.append("В документе нет товарных позиций с кодами маркировки для Excel 1С.")
    if errors:
        raise OneCExportValidationError(errors)

    items = []
    for data in sorted(grouped.values(), key=lambda item: item["position_number"]):
        items.append(
            OneCExportItem(
                article=data["article"],
                honest_sign_name=data["honest_sign_name"],
                gender=data["gender"],
                country_of_origin=data["country_of_origin"],
                brand=data["brand"],
                manufacturer=data["manufacturer"],
                composition=data["composition"],
                size=data["size"],
                color=data["color"],
                retail_price=data["retail_price"],
                marking_code_count=len(data["codes"]),
                ean13=data["ean13"],
                category=data["category"],
                gtin=data["gtin"],
            )
        )
    return items


def copy_row_style(source_cells, target_cells):
    for source, target in zip(source_cells, target_cells):
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def add_template_validations(worksheet, last_row):
    validations = [
        ("'Пол список'!$B$2:$B$6", f"D4:D{last_row}", "Пол - Выберите пол из списка"),
        (
            "'Классификатор стран'!$B$2:$B$252",
            f"E4:E{last_row}",
            "Страна - Выберите страну из списка",
        ),
        (
            "Категория!$B$2:$B$5",
            f"N4:N{last_row}",
            "Категория - Выберите категорию из списка",
        ),
    ]
    for formula, cell_range, prompt in validations:
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.promptTitle = "Выбор значения"
        validation.prompt = prompt
        validation.showInputMessage = True
        validation.showErrorMessage = True
        worksheet.add_data_validation(validation)
        validation.add(cell_range)


def render_one_c_xlsx(items, output_path, template_path=MARKING_ONE_C_TEMPLATE_PATH, consignor=MARKING_ONE_C_CONSIGNOR):
    template_path = Path(template_path)
    if not template_path.exists():
        raise RuntimeError(f"Не найден шаблон Excel для 1С: {template_path}")

    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Data Validation extension is not supported")
        workbook = load_workbook(template_path)
    worksheet = workbook[ONE_C_SHEET_NAME]
    actual_headers = [worksheet.cell(3, column).value for column in range(1, 15)]
    if actual_headers != ONE_C_HEADERS:
        raise RuntimeError("Структура шаблона Excel для 1С не соответствует ожидаемым 14 столбцам.")

    style_cells = [copy(worksheet.cell(4, column)) for column in range(1, 15)]
    template_row_height = worksheet.row_dimensions[4].height
    if worksheet.max_row >= 4:
        worksheet.delete_rows(4, worksheet.max_row - 3)
    if consignor:
        worksheet["A1"] = consignor

    for row_number, item in enumerate(items, start=4):
        values = [
            row_number - 3,
            item.article,
            item.honest_sign_name,
            item.gender,
            item.country_of_origin,
            item.brand,
            item.manufacturer,
            item.composition,
            item.size,
            item.color,
            float(item.retail_price),
            item.marking_code_count,
            item.ean13,
            item.category,
        ]
        target_cells = [worksheet.cell(row_number, column) for column in range(1, 15)]
        copy_row_style(style_cells, target_cells)
        for cell, value in zip(target_cells, values):
            cell.value = value
        worksheet.row_dimensions[row_number].height = template_row_height
        worksheet.cell(row_number, 2).number_format = "@"
        worksheet.cell(row_number, 11).number_format = "0.00"
        worksheet.cell(row_number, 12).number_format = "0"
        worksheet.cell(row_number, 13).number_format = "@"

    add_template_validations(worksheet, max(4, len(items) + 3))
    output_path = Path(output_path)
    workbook.save(output_path)
    return output_path


def create_one_c_xlsx(rows, catalog_names, output_path):
    items = build_one_c_export_items(rows, catalog_names)
    render_one_c_xlsx(items, output_path)
    return output_path, items
