from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import MOYSKLAD_API_BASE_URL, MOYSKLAD_CA_FILE, MOYSKLAD_SSL_VERIFY, MOYSKLAD_TOKEN
from modules.moysklad.client import MoySkladClient, MoySkladError


HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")


def parse_bool(value, default=True):
    if value is None:
        return default

    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "y", "on", "да"}:
        return True
    if normalized in {"0", "false", "no", "n", "off", "нет"}:
        return False
    return default


def build_moysklad_client():
    return MoySkladClient(
        token=MOYSKLAD_TOKEN,
        base_url=MOYSKLAD_API_BASE_URL,
        ssl_verify=parse_bool(MOYSKLAD_SSL_VERIFY),
        ca_file=MOYSKLAD_CA_FILE or None,
    )


def response_rows(payload):
    return payload.get("rows", []) if isinstance(payload, dict) else []


def unique(values):
    result = []
    seen = set()
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def extract_cis_codes(value):
    codes = []

    def walk(node):
        if isinstance(node, dict):
            cis = node.get("cis")
            if cis:
                codes.append(str(cis))

            for child in node.get("trackingCodes") or []:
                walk(child)
        elif isinstance(node, list):
            for item in node:
                walk(item)

    walk(value)
    return unique(codes)


def find_retireorder(client, document_name):
    documents_by_id = {}
    attempts = [
        {"filter": f"name={document_name}", "limit": 10},
        {"search": document_name, "limit": 10},
    ]

    for params in attempts:
        payload = client.list_entities("retireorder", params=params)
        for document in response_rows(payload):
            document_id = str(document.get("id") or "").strip()
            if document_id:
                documents_by_id[document_id] = document

    exact_matches = [
        document
        for document in documents_by_id.values()
        if str(document.get("name") or "").strip() == document_name
    ]
    matches = exact_matches or list(documents_by_id.values())

    if not matches:
        raise MoySkladError(f"Документ вывода из оборота «{document_name}» не найден.")
    if len(matches) > 1:
        names = ", ".join(str(document.get("name") or document.get("id")) for document in matches[:5])
        raise MoySkladError(f"Найдено несколько документов: {names}. Уточните название документа.")

    return matches[0]


def get_assortment_name(position):
    assortment = position.get("assortment") or {}
    name = str(assortment.get("name") or "").strip()
    if name:
        return name

    meta = assortment.get("meta") or {}
    href = str(meta.get("href") or "").strip()
    return href.rsplit("/", 1)[-1] if href else "Без названия"


def get_all_position_tracking_codes(client, entity_type, entity_id, position_id):
    codes = []
    limit = 100
    offset = 0

    while True:
        payload = client.get_position_tracking_codes(
            entity_type,
            entity_id,
            position_id,
            params={"codetype": "gs1", "limit": limit, "offset": offset},
        )
        rows = response_rows(payload)
        codes.extend(extract_cis_codes(rows))

        meta = payload.get("meta") or {}
        size = int(meta.get("size") or 0)
        offset += limit
        if offset >= size or not rows:
            break

    return unique(codes)


def get_retireorder_export_rows(client, document_name):
    document = find_retireorder(client, document_name)
    document_id = document["id"]
    positions_payload = client.get_positions(
        "retireorder",
        document_id,
        params={"expand": "assortment", "limit": 1000},
    )

    result = []
    for position in response_rows(positions_payload):
        position_id = str(position.get("id") or "").strip()
        if not position_id:
            continue

        codes = get_all_position_tracking_codes(client, "retireorder", document_id, position_id)
        result.append(
            {
                "name": get_assortment_name(position),
                "quantity": position.get("quantity"),
                "codes": codes,
            }
        )

    return document, result


def create_trend_island_codes_xlsx(rows, output_path):
    output_path = Path(output_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Коды маркировки"

    headers = ["Название", "Количество", "Коды маркировки"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_index = 2
    for item in rows:
        codes = item["codes"] or [""]
        for code_index, code in enumerate(codes):
            if code_index == 0:
                ws.cell(row=row_index, column=1, value=item["name"])
                ws.cell(row=row_index, column=2, value=item["quantity"])

            ws.cell(row=row_index, column=3, value=code)
            for col in range(1, 4):
                ws.cell(row=row_index, column=col).alignment = Alignment(
                    vertical="top",
                    wrap_text=col in (1, 3),
                )
            ws.row_dimensions[row_index].height = 24
            row_index += 1

    widths = [42, 14, 52]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    return output_path
