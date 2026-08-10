import html
import logging
import re
from datetime import datetime
from io import BytesIO

from telegram import InlineKeyboardButton, InlineKeyboardMarkup, InputFile, InputMediaPhoto, Update
from telegram.constants import ParseMode
from telegram.ext import CallbackQueryHandler, ContextTypes, ConversationHandler, MessageHandler, filters
from openpyxl import load_workbook

from config import GROUP_CHAT_ID, LAMODA_RETURNS_TOPIC_ID
from modules.lamoda_fbs.constants import MANAGER_ROLES, PackState
from modules.lamoda_fbs.keyboards import (
    assembly_confirm,
    back_menu,
    batch_actions,
    cargo_menu_keyboard,
    cargo_scan_keyboard,
    defect_photos,
    main_menu,
    marking_menu,
    return_condition,
    returns_menu,
)
from modules.lamoda_fbs.marking import MarkingValidationError, parse_marking_code, short_kiz, validate_against_catalog
from modules.lamoda_fbs.services import (
    assembly_label_documents,
    create_lamoda_shipment,
    create_marking_documents,
    discover_orders,
    get_client,
    order_summary,
    prepare_orders,
    resolve_return_barcode,
    resolve_return_order,
    shipment_documents,
)
from modules.lamoda_fbs.storage import (
    Shipment,
    add_pack_to_cargo,
    assign_marking_code,
    cancel_marking_batch,
    cargo_manifest,
    code_fingerprint,
    complete_pack_without_marking,
    confirm_marking_batch,
    create_cargo_place,
    get_active_session,
    get_next_unscanned_pack,
    get_session_packs,
    list_expected_returns,
    list_problem_packs,
    list_return_receipts,
    mark_pack_barcode_scanned,
    marking_batch_rows,
    marking_history,
    pending_counts,
    record_return_receipt,
    set_cargo_status,
    set_session_labels_ready,
)
from modules.employees.roles import has_any_role
from modules.payroll.google_sheets import find_employee_for_telegram_user, get_employees
from modules.storage.postgres import session_scope


logger = logging.getLogger(__name__)

SCAN_PACK, SCAN_KIZ, CARGO_SCAN, RETURN_PHOTO, RETURN_BARCODE, RETURN_MANUAL_ORDER, RETURN_KIZ, RETURN_CONDITION, RETURN_DEFECT_REASON, RETURN_DEFECT_PHOTOS, BATCH_ERRORS = range(11)


def _employee(user):
    try:
        return find_employee_for_telegram_user(user) or {}
    except Exception:
        logger.exception("Could not resolve Lamoda employee")
        return {}


def _employee_name(user):
    employee = _employee(user)
    return employee.get("full_name") or user.full_name or str(user.id)


def _is_lamoda_manager(user):
    return has_any_role(_employee(user), MANAGER_ROLES)


async def _answer_error(update, error):
    logger.exception("Lamoda operation failed")
    text = f"⚠️ {error}"
    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.message.reply_text(text, reply_markup=back_menu())
    else:
        await update.message.reply_text(text)


async def show_lamoda_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    context.user_data.clear()
    await query.edit_message_text("🛒 Lamoda FBS\n\nВыберите действие:", reply_markup=main_menu())
    return ConversationHandler.END


async def assembly_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("Загружаю новые FBS-заказы Lamoda…")
    try:
        orders = await discover_orders()
        summary = order_summary(orders)
        context.user_data["lamoda_order_ids"] = [str(row.get("id") or row.get("orderId")) for row in orders]
        nearest = summary["nearest_cutoff"]
        cutoff_text = nearest.astimezone().strftime("%d.%m.%Y %H:%M") if nearest else "—"
        await query.edit_message_text(
            f"Найдено заказов: {summary['orders']}\n"
            f"Товаров: {summary['items']}\n"
            f"Ближайший cutOff: {cutoff_text}",
            reply_markup=assembly_confirm() if orders else back_menu(),
        )
    except Exception as error:
        await _answer_error(update, error)


async def assembly_prepare(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("Создаю сборки и отдельную упаковку для каждого товара…")
    try:
        client = get_client()
        order_ids = context.user_data.get("lamoda_order_ids") or []
        orders = [await client.get_order(order_id) for order_id in order_ids]
        result = await prepare_orders(orders, update.effective_user.id, _employee_name(update.effective_user), client)
        if not get_session_packs(result["session_id"]):
            error_text = "; ".join(f"{row['order_id']}: {row['error']}" for row in result["errors"])
            await query.message.reply_text(
                f"⚠️ Ни один заказ не подготовлен.\n{error_text}",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("🔄 Повторить неуспешные заказы", callback_data="lamoda:assembly:start")],
                    [InlineKeyboardButton("⬅️ Lamoda FBS", callback_data="section:lamoda")],
                ]),
            )
            return
        try:
            documents = await assembly_label_documents(result["session_id"], client)
        except Exception as error:
            await query.message.reply_text(
                f"⚠️ Заказы подготовлены, но этикетки получить не удалось: {error}",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("🔄 Повторить получение этикеток", callback_data=f"lamoda:labels:retry:{result['session_id']}")],
                    [InlineKeyboardButton("⬅️ Lamoda FBS", callback_data="section:lamoda")],
                ]),
            )
            return
        await _send_assembly_label_docs(query.message, result["session_id"], documents)
        labels_ready = not documents["excluded_items"] and not documents["excluded_packs"]
        set_session_labels_ready(result["session_id"], labels_ready)
        warnings = []
        if result["errors"]:
            warnings.append("Ошибки: " + "; ".join(f"{row['order_id']}: {row['error']}" for row in result["errors"]))
        if documents["excluded_items"]:
            warnings.append("Исключены item: " + ", ".join(documents["excluded_items"]))
        if documents["excluded_packs"]:
            warnings.append("Исключены pack: " + ", ".join(documents["excluded_packs"]))
        action_rows = []
        if documents["excluded_items"] or documents["excluded_packs"]:
            action_rows.append([InlineKeyboardButton("🔄 Повторить получение этикеток", callback_data=f"lamoda:labels:retry:{result['session_id']}")])
        if result["errors"]:
            action_rows.append([InlineKeyboardButton("🔄 Повторить неуспешные заказы", callback_data="lamoda:assembly:start")])
        if labels_ready:
            action_rows.append([InlineKeyboardButton("▶️ Перейти к сборке", callback_data="lamoda:assembly:continue")])
        action_rows.append([InlineKeyboardButton("⬅️ Lamoda FBS", callback_data="section:lamoda")])
        await query.message.reply_text(
            f"✅ Подготовлено заказов: {len(result['successes'])}.\n"
            + ("\n⚠️ " + "\n⚠️ ".join(warnings) if warnings else ""),
            reply_markup=InlineKeyboardMarkup(action_rows),
        )
    except Exception as error:
        await _answer_error(update, error)


async def _send_assembly_label_docs(message, session_id, documents):
    await message.reply_document(
        InputFile(BytesIO(documents["item_pdf"]), filename=f"lamoda_item_labels_{session_id}.pdf"),
        caption="🏷 Товарные этикетки Lamoda · формат S (58×40 мм)",
    )
    await message.reply_document(
        InputFile(BytesIO(documents["pack_pdf"]), filename=f"lamoda_pack_labels_{session_id}.pdf"),
        caption="📦 Паковые этикетки Lamoda · формат M (75×120 мм)",
    )


async def assembly_labels_retry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    session_id = int(query.data.rsplit(":", 1)[1])
    try:
        documents = await assembly_label_documents(session_id)
        await _send_assembly_label_docs(query.message, session_id, documents)
        excluded = documents["excluded_items"] + documents["excluded_packs"]
        set_session_labels_ready(session_id, not excluded)
        rows = []
        if excluded:
            rows.append([InlineKeyboardButton("🔄 Повторить ещё раз", callback_data=f"lamoda:labels:retry:{session_id}")])
        else:
            rows.append([InlineKeyboardButton("▶️ Перейти к сборке", callback_data="lamoda:assembly:continue")])
        rows.append([InlineKeyboardButton("⬅️ Lamoda FBS", callback_data="section:lamoda")])
        await query.message.reply_text(
            "✅ Этикетки получены повторно."
            + (f"\n⚠️ Всё ещё исключены: {', '.join(excluded)}" if excluded else ""),
            reply_markup=InlineKeyboardMarkup(rows),
        )
    except Exception as error:
        await _answer_error(update, error)


def _pack_prompt(pack, position, total):
    return (
        f"📦 Товар {position} из {total}\n\n"
        f"Заказ: {pack['order_id']}\n"
        f"Название: {pack['product_name'] or '—'}\n"
        f"Размер: {pack['size'] or '—'}\n"
        f"Артикул: {pack['sku'] or '—'}\n"
        f"Товарная этикетка / itemId: {pack['item_id']}\n"
        f"Паковая этикетка / packNumber: {pack['pack_number']}\n\n"
        "1. Отсканируйте паковую этикетку."
    )


async def _send_next_pack(update, context):
    session_id = context.user_data["lamoda_session_id"]
    pack = get_next_unscanned_pack(session_id)
    if not pack:
        await update.effective_message.reply_text(
            "✅ Все товары собраны. Переходите к созданию грузовых мест.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🚚 Грузовые места", callback_data="lamoda:cargo:menu")]]),
        )
        return ConversationHandler.END
    packs = get_session_packs(session_id)
    completed = sum(1 for row in packs if row["packed"])
    context.user_data["lamoda_expected_pack"] = pack
    await update.effective_message.reply_text(_pack_prompt(pack, completed + 1, len(packs)))
    return SCAN_PACK


async def assembly_continue(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    active = get_active_session()
    if not active:
        await query.message.reply_text("Активной сборки нет.", reply_markup=back_menu())
        return ConversationHandler.END
    if not active.labels_ready:
        await query.message.reply_text(
            "⚠️ Комплект этикеток ещё не готов. Повторите получение этикеток.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔄 Получить этикетки", callback_data=f"lamoda:labels:retry:{active.id}")]
            ]),
        )
        return ConversationHandler.END
    context.user_data["lamoda_session_id"] = active.id
    return await _send_next_pack(update, context)


async def scan_pack(update: Update, context: ContextTypes.DEFAULT_TYPE):
    expected = context.user_data.get("lamoda_expected_pack")
    try:
        mark_pack_barcode_scanned(update.message.text, expected["pack_number"])
    except Exception as error:
        await update.message.reply_text(f"⚠️ {error}\nПовторите сканирование packNumber.")
        return SCAN_PACK
    await update.message.reply_text(
        "2. Отсканируйте полный КИЗ товара или укажите, что товар немаркируемый.",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(
            "📦 Без КИЗ — товар немаркируемый",
            callback_data=f"lamoda:kiz:skip:{expected['pack_number']}",
        )]]),
    )
    return SCAN_KIZ


async def scan_kiz_skip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    expected = context.user_data.get("lamoda_expected_pack") or {}
    pack_number = query.data.removeprefix("lamoda:kiz:skip:")
    if pack_number != expected.get("pack_number"):
        await query.answer("Это кнопка от другого товара.", show_alert=True)
        return SCAN_KIZ
    await query.answer()
    try:
        complete_pack_without_marking(pack_number, update.effective_user.id)
    except Exception as error:
        await query.message.reply_text(f"⚠️ {error}")
        return SCAN_KIZ
    await query.edit_message_text("✅ Немаркируемый товар собран без КИЗ.")
    return await _send_next_pack(update, context)


async def scan_kiz(update: Update, context: ContextTypes.DEFAULT_TYPE):
    pack = context.user_data["lamoda_expected_pack"]
    try:
        parsed = parse_marking_code(update.message.text)
        warnings = validate_against_catalog(parsed, pack["product_name"], pack["size"])
        assign_marking_code(
            pack["pack_number"], parsed.raw, parsed.uit, parsed.gtin, parsed.serial,
            update.effective_user.id,
        )
        await update.message.reply_text("✅ Товар собран." + ("\n⚠️ " + " ".join(warnings) if warnings else ""))
    except Exception as error:
        await update.message.reply_text(f"⚠️ {error}\nОтсканируйте КИЗ ещё раз.")
        return SCAN_KIZ
    return await _send_next_pack(update, context)


def _manifest_text(manifest):
    if not manifest:
        return "Грузовые места ещё не созданы."
    return "\n".join(
        f"📦 №{row['local_number']}: {len(row['packs'])} уп. · {row['status']}"
        for row in manifest
    )


async def cargo_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    active = get_active_session()
    if not active:
        await query.edit_message_text("Активной сборки нет.", reply_markup=back_menu())
        return
    manifest = cargo_manifest(active.id)
    packs = [pack for cargo in manifest for pack in cargo["packs"]]
    summary = (
        f"\n\nЗаказов: {len({pack['order_id'] for pack in packs})}\n"
        f"Товаров: {len(packs)}\nУпаковок: {len(packs)}\n"
        f"Грузовых мест: {len(manifest)}"
    )
    await query.edit_message_text(
        f"🚚 Грузовые места\n\n{_manifest_text(manifest)}{summary}",
        reply_markup=cargo_menu_keyboard(active.id, manifest),
    )


async def cargo_new(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    session_id = int(query.data.rsplit(":", 1)[1])
    try:
        cargo = create_cargo_place(session_id, update.effective_user.id)
        context.user_data["lamoda_cargo_id"] = cargo.id
        await query.message.reply_text(
            f"📦 Грузовое место №{cargo.local_number}\n\n"
            f"Пометьте коробку цифрой «{cargo.local_number}» и сканируйте в бот паковые этикетки.",
            reply_markup=cargo_scan_keyboard(cargo.id),
        )
        return CARGO_SCAN
    except Exception as error:
        await _answer_error(update, error)
        return ConversationHandler.END


async def cargo_scan(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cargo_id = context.user_data["lamoda_cargo_id"]
    try:
        added = add_pack_to_cargo(cargo_id, update.message.text, update.effective_user.id)
        text = "добавлена" if added else "уже была добавлена"
        await update.message.reply_text(f"✅ Упаковка {update.message.text.strip()} {text}.", reply_markup=cargo_scan_keyboard(cargo_id))
    except Exception as error:
        await update.message.reply_text(f"⚠️ {error}", reply_markup=cargo_scan_keyboard(cargo_id))
    return CARGO_SCAN


async def cargo_close(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    cargo_id = int(query.data.rsplit(":", 1)[1])
    try:
        set_cargo_status(cargo_id, "CLOSED", update.effective_user.id)
        await query.message.reply_text("✅ Грузовое место закрыто.", reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("➕ Следующее грузовое место", callback_data="lamoda:cargo:menu")]
        ]))
    except Exception as error:
        await query.message.reply_text(f"⚠️ {error}")
        return CARGO_SCAN
    return ConversationHandler.END


async def cargo_reopen(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    cargo_id = int(query.data.rsplit(":", 1)[1])
    try:
        set_cargo_status(cargo_id, "OPEN", update.effective_user.id)
        context.user_data["lamoda_cargo_id"] = cargo_id
        await query.message.reply_text("🔓 Место снова открыто. Сканируйте упаковки.", reply_markup=cargo_scan_keyboard(cargo_id))
        return CARGO_SCAN
    except Exception as error:
        await _answer_error(update, error)
        return ConversationHandler.END


async def shipment_create(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    session_id = int(query.data.rsplit(":", 1)[1])
    await query.message.reply_text("Создаю одну общую отгрузку Lamoda…")
    try:
        shipment_id = await create_lamoda_shipment(session_id)
        await _send_shipment_documents(query.message, session_id)
        await query.message.reply_text(f"✅ Отгрузка {shipment_id} создана.")
    except Exception as error:
        await _answer_error(update, error)


async def _send_shipment_documents(message, session_id):
    documents = await shipment_documents(session_id)
    for document in documents["pallet_documents"]:
        number = document["local_number"]
        await message.reply_document(
            InputFile(BytesIO(document["content"]), filename=f"lamoda_грузовое_место_{number:02d}.pdf"),
            caption=(f"📦 Грузовое место №{number}\nКод Lamoda: {document['pallet_id']}\n"
                     f"Упаковок: {document['pack_count']}\n\nНаклейте этикетку на короб, ранее помеченный цифрой «{number}»."),
        )
    await message.reply_document(
        InputFile(BytesIO(documents["manifest_pdf"]), filename=f"lamoda_shipment_manifest_{documents['shipment_id']}.pdf"),
        caption="📋 Состав грузовых мест",
    )


async def shipments_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    with session_scope() as session:
        rows = session.query(Shipment).order_by(Shipment.id.desc()).limit(10).all()
    keyboard = [[InlineKeyboardButton(f"📄 {row.shipment_id}", callback_data=f"lamoda:shipment:docs:{row.session_id}")] for row in rows]
    keyboard.append([InlineKeyboardButton("⬅️ Назад", callback_data="section:lamoda")])
    await query.edit_message_text("📤 Последние отгрузки:" if rows else "Отгрузок ещё нет.", reply_markup=InlineKeyboardMarkup(keyboard))


async def shipment_docs_repeat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    try:
        await _send_shipment_documents(query.message, int(query.data.rsplit(":", 1)[1]))
    except Exception as error:
        await _answer_error(update, error)


async def returns_open(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("↩️ Возвраты Lamoda", reply_markup=returns_menu())


async def return_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    context.user_data["lamoda_return"] = {}
    await query.message.reply_text("📷 Сфотографируйте этикетку возвратного отправления.")
    return RETURN_PHOTO


async def return_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["lamoda_return"]["label_photo"] = update.message.photo[-1].file_id
    await update.message.reply_text("🔎 Отсканируйте packNumber / штрихкод этикетки.")
    return RETURN_BARCODE


async def return_barcode(update: Update, context: ContextTypes.DEFAULT_TYPE):
    data = context.user_data["lamoda_return"]
    data["barcode"] = update.message.text.strip()
    await update.message.reply_text("Ищу возврат в базе и Lamoda…")
    try:
        pack = await resolve_return_barcode(data["barcode"])
    except Exception as error:
        logger.exception("Return lookup failed")
        pack = None
        data["lookup_error"] = str(error)
    if not pack:
        data["problematic"] = True
        data["problem_reason"] = "Возврат не найден автоматически."
        await update.message.reply_text("⚠️ Отправление не найдено. Введите номер заказа вручную; приёмка будет отмечена как проблемная.")
        return RETURN_MANUAL_ORDER
    data["pack"] = pack
    data["problematic"] = False
    data["problem_reason"] = ""
    await update.message.reply_text(
        f"Заказ: {pack['order_id']}\npackNumber: {pack['pack_number']}\nitemId: {pack['item_id']}\n"
        f"Товар: {pack['product_name'] or '—'}\nРазмер: {pack['size'] or '—'}\nАртикул: {pack['sku'] or '—'}\n"
        f"КИЗ: {short_kiz(pack['raw_code'])}\nСтатус Lamoda: {pack.get('return_status') or pack['lamoda_status'] or '—'}\n"
        f"Тип возврата: {pack.get('return_type') or '—'}\n"
        f"Дата возврата: {pack.get('return_date') or '—'}\n\n"
        "Повторно отсканируйте КИЗ товара."
    )
    return RETURN_KIZ


async def return_manual_order(update: Update, context: ContextTypes.DEFAULT_TYPE):
    data = context.user_data["lamoda_return"]
    data["manual_order_id"] = update.message.text.strip()
    try:
        candidates = await resolve_return_order(data["manual_order_id"])
    except Exception:
        logger.exception("Manual Lamoda return order lookup failed")
        candidates = []
    if len(candidates) == 1:
        data["pack"] = candidates[0]
        data["problematic"] = False
        data["problem_reason"] = ""
        await update.message.reply_text(
            f"Найден pack {candidates[0]['pack_number']} · {candidates[0]['product_name'] or 'без названия'}.\n"
            "Отсканируйте КИЗ товара."
        )
        return RETURN_KIZ
    if len(candidates) > 1:
        variants = "\n".join(f"• {row['pack_number']} · {row['product_name']} · {row['size']}" for row in candidates)
        await update.message.reply_text(
            f"В заказе несколько упаковок:\n{variants}\n\nОтсканируйте нужный packNumber."
        )
        return RETURN_BARCODE
    data["pack"] = {
        "pack_number": data["barcode"], "order_id": data["manual_order_id"], "item_id": "",
        "product_name": "Неопознанный товар", "size": "", "sku": "", "raw_code": "",
        "fingerprint": "", "lamoda_status": "",
    }
    await update.message.reply_text("Отсканируйте КИЗ товара.")
    return RETURN_KIZ


async def return_kiz(update: Update, context: ContextTypes.DEFAULT_TYPE):
    data = context.user_data["lamoda_return"]
    try:
        parsed = parse_marking_code(update.message.text)
    except MarkingValidationError as error:
        await update.message.reply_text(f"⚠️ {error}\nПовторите сканирование.")
        return RETURN_KIZ
    data["scanned_kiz_fingerprint"] = code_fingerprint(parsed.normalized)
    original = data["pack"].get("fingerprint")
    data["kiz_matches"] = bool(original and original == data["scanned_kiz_fingerprint"])
    if not data["kiz_matches"]:
        data["problematic"] = True
        data["problem_reason"] = (data.get("problem_reason", "") + " КИЗ не совпадает с исходным.").strip()
    await update.message.reply_text(
        ("✅ КИЗ совпадает.\n" if data["kiz_matches"] else "⚠️ КИЗ не совпадает; приёмка будет проблемной.\n")
        + "Укажите состояние товара:", reply_markup=return_condition(),
    )
    return RETURN_CONDITION


async def return_condition_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    condition = query.data.rsplit(":", 1)[1]
    context.user_data["lamoda_return"]["condition"] = condition
    if condition == "NORMAL":
        return await _save_return(update, context)
    await query.message.reply_text("Опишите причину брака.")
    return RETURN_DEFECT_REASON


async def return_defect_reason(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["lamoda_return"]["defect_reason"] = update.message.text.strip()
    context.user_data["lamoda_return"]["defect_photos"] = []
    await update.message.reply_text("📷 Пришлите минимум одну фотографию брака, затем нажмите «Готово».", reply_markup=defect_photos())
    return RETURN_DEFECT_PHOTOS


async def return_defect_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    photos = context.user_data["lamoda_return"].setdefault("defect_photos", [])
    if len(photos) >= 9:
        await update.message.reply_text("Можно приложить до 9 фотографий брака.", reply_markup=defect_photos())
        return RETURN_DEFECT_PHOTOS
    photos.append(update.message.photo[-1].file_id)
    await update.message.reply_text(f"Фото добавлено. Всего: {len(photos)}.", reply_markup=defect_photos())
    return RETURN_DEFECT_PHOTOS


async def return_defect_done(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not context.user_data["lamoda_return"].get("defect_photos"):
        await query.answer("Добавьте хотя бы одно фото брака.", show_alert=True)
        return RETURN_DEFECT_PHOTOS
    await query.answer()
    return await _save_return(update, context)


def _mentions_for_roles(required_roles):
    mentions = []
    try:
        employees = get_employees(include_inactive=False)
    except Exception:
        logger.exception("Could not load employees for Lamoda return mention")
        employees = []
    seen = set()
    for employee in employees:
        if not has_any_role(employee, required_roles):
            continue
        identity = str(employee.get("employee_id") or employee.get("telegram_user_id") or "").strip()
        if identity and identity in seen:
            continue
        if identity:
            seen.add(identity)
        username = str(employee.get("telegram_username") or "").strip().lstrip("@")
        name = html.escape(employee.get("full_name") or "Сотрудник")
        if username:
            mentions.append(f"@{html.escape(username)}")
        elif employee.get("telegram_user_id"):
            mentions.append(f'<a href="tg://user?id={employee["telegram_user_id"]}">{name}</a>')
    return " ".join(mentions)


async def _save_return(update, context):
    data = context.user_data["lamoda_return"]
    data["employee_name"] = _employee_name(update.effective_user)
    pack = data["pack"]
    status_upper = str(pack.get("lamoda_status") or "").upper()
    if pack.get("fingerprint") and not any(token in status_upper for token in ("RETURN", "NOT_BOUGHT", "NOTBOUGHT")):
        data["problematic"] = True
        data["problem_reason"] = (data.get("problem_reason", "") + " В Lamoda нет возвратного статуса.").strip()
    try:
        receipt_id = record_return_receipt(
            pack_number=pack.get("pack_number"), order_id=pack.get("order_id"), item_id=pack.get("item_id"),
            return_item_id=pack.get("return_item_id", ""), condition=data["condition"],
            defect_reason=data.get("defect_reason", ""), label_photo_file_id=data["label_photo"],
            scanned_kiz_fingerprint=data.get("scanned_kiz_fingerprint", ""),
            defect_photo_file_ids=data.get("defect_photos", []), user_id=update.effective_user.id,
            user_name=_employee_name(update.effective_user), problematic=data.get("problematic", False),
            problem_reason=data.get("problem_reason", ""), manager_override=_is_lamoda_manager(update.effective_user),
        )
    except Exception as error:
        data["problematic"] = True
        data["problem_reason"] = (data.get("problem_reason", "") + f" {error}").strip()
        try:
            await _send_return_report(context, data, None)
        except Exception:
            logger.exception("Could not notify managers about failed Lamoda return receipt")
        await update.effective_message.reply_text(f"⚠️ {error}", reply_markup=returns_menu())
        return ConversationHandler.END
    report_warning = ""
    try:
        await _send_return_report(context, data, receipt_id)
    except Exception:
        logger.exception("Lamoda return receipt was saved but topic report failed: receipt_id=%s", receipt_id)
        report_warning = "\n⚠️ Отчёт в Telegram-тему не отправлен; сама приёмка сохранена."
    await update.effective_message.reply_text(
        f"✅ Возврат принят. Запись №{receipt_id}.{report_warning}",
        reply_markup=returns_menu(),
    )
    context.user_data.pop("lamoda_return", None)
    return ConversationHandler.END


async def _send_return_report(context, data, receipt_id):
    if not GROUP_CHAT_ID or not LAMODA_RETURNS_TOPIC_ID:
        logger.warning("Lamoda return topic is not configured; receipt_id=%s", receipt_id)
        return
    pack = data["pack"]
    defect = data["condition"] == "DEFECT"
    attention = defect or data.get("problematic")
    heading = "❌ Принят бракованный возврат Lamoda" if defect else "✅ Принят возврат Lamoda"
    caption = (
        f"{heading}\n\nСотрудник: {html.escape(data.get('employee_name') or '')}\n"
        f"Заказ: {html.escape(str(pack.get('order_id') or '—'))}\nПак: {html.escape(str(pack.get('pack_number') or '—'))}\n"
        f"Товар: {html.escape(pack.get('product_name') or '—')}\nРазмер: {html.escape(pack.get('size') or '—')}\n"
        f"Статус: {'Брак' if defect else 'Норм'}\n"
        + (f"Причина: {html.escape((data.get('defect_reason') or '')[:500])}\n" if defect else "")
        + f"КИЗ совпадает: {'да' if data.get('kiz_matches') else 'нет'}\n"
        + f"Маркировка: {'требует сверки' if data.get('problematic') else 'ожидает возврата в оборот'}\n"
        + (f"⚠️ {html.escape((data.get('problem_reason') or '')[:300])}\n" if data.get("problematic") else "")
        + f"Дата: {datetime.now():%d.%m.%Y %H:%M}"
    )
    if attention:
        mention_roles = (
            {"brand_manager", "warehouse_manager", "operations"}
            if defect
            else {"warehouse_manager"}
        )
        mentions = _mentions_for_roles(mention_roles)
        if mentions:
            caption += "\n\n" + mentions
    chat_id = int(GROUP_CHAT_ID)
    thread_id = int(LAMODA_RETURNS_TOPIC_ID)
    photos = [data["label_photo"]] + data.get("defect_photos", [])
    if len(photos) == 1:
        await context.bot.send_photo(chat_id=chat_id, message_thread_id=thread_id, photo=photos[0], caption=caption, parse_mode=ParseMode.HTML)
    else:
        media = [InputMediaPhoto(media=file_id, caption=caption if index == 0 else None, parse_mode=ParseMode.HTML if index == 0 else None) for index, file_id in enumerate(photos)]
        await context.bot.send_media_group(chat_id=chat_id, message_thread_id=thread_id, media=media)


async def returns_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    mode = query.data.rsplit(":", 1)[1]
    if mode == "problems" and not _is_lamoda_manager(update.effective_user):
        await query.answer("Проблемные операции доступны только руководителю.", show_alert=True)
        return
    await query.answer()
    if mode == "expected":
        rows = list_expected_returns()
        lines = [f"• {row['pack_number']} · {row['order_id']} · {row['product_name']}" for row in rows]
        title = "📦 Ожидаемые возвраты"
    else:
        if mode == "problems":
            problem_packs = list_problem_packs()
            problem_receipts = list_return_receipts(problematic_only=True)
            lines = [f"• pack {row['pack_number']} · {row['order_id']} · {row['product_name']}" for row in problem_packs]
            lines.extend(f"• приёмка №{row['id']} · {row['pack_number']} · {row['problem_reason']}" for row in problem_receipts)
            title = "⚠️ Проблемные операции Lamoda"
        else:
            rows = list_return_receipts()
            lines = [f"• №{row['id']} · {row['pack_number']} · {row['condition']} · {row['created_at']:%d.%m %H:%M}" for row in rows]
            title = "📋 Последние приёмки"
    await query.edit_message_text(title + "\n\n" + ("\n".join(lines) if lines else "Записей нет."), reply_markup=returns_menu())


async def marking_open(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not _is_lamoda_manager(update.effective_user):
        await query.answer("Раздел доступен только руководителю склада.", show_alert=True)
        return
    await query.answer()
    await query.edit_message_text("🏷 Операции ЧЗ Lamoda", reply_markup=marking_menu(pending_counts()))


async def marking_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not _is_lamoda_manager(update.effective_user):
        await query.answer("Нет доступа.", show_alert=True)
        return
    await query.answer()
    batch_type = query.data.rsplit(":", 1)[1]
    try:
        docs = create_marking_documents(batch_type, update.effective_user.id, _employee_name(update.effective_user))
        batch_id = docs["batch_id"]
        await query.message.reply_document(InputFile(BytesIO(docs["pdf"]), filename=f"lamoda_marking_{batch_id}.pdf"))
        await query.message.reply_document(InputFile(BytesIO(docs["xlsx"]), filename=f"lamoda_marking_{batch_id}.xlsx"))
        await query.message.reply_text(
            f"Партия №{batch_id} · кодов: {len(docs['rows'])}.\nПодтвердите итог операции в Честном ЗНАКе:",
            reply_markup=batch_actions(batch_id, batch_type),
        )
    except Exception as error:
        await _answer_error(update, error)


async def batch_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if not _is_lamoda_manager(update.effective_user):
        return
    batch_id = int(query.data.rsplit(":", 1)[1])
    try:
        status = confirm_marking_batch(batch_id, update.effective_user.id, manager_name=_employee_name(update.effective_user))
        await query.edit_message_text(f"✅ Партия №{batch_id} подтверждена. Статус: {status}.", reply_markup=back_menu())
    except Exception as error:
        await _answer_error(update, error)


async def batch_errors_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if not _is_lamoda_manager(update.effective_user):
        return ConversationHandler.END
    context.user_data["lamoda_batch_id"] = int(query.data.rsplit(":", 1)[1])
    await query.message.reply_text(
        "Введите packNumber или полные КИЗ неуспешных позиций "
        "через запятую/с новой строки или пришлите TXT, CSV, XLSX."
    )
    return BATCH_ERRORS


async def batch_errors_received(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        failed = _resolve_failed_entries(
            context.user_data["lamoda_batch_id"],
            [value for value in re.split(r"[,;\s]+", update.message.text) if value],
        )
    except Exception as error:
        await update.message.reply_text(f"⚠️ {error}")
        return BATCH_ERRORS
    return await _confirm_batch_errors(update, context, failed)


def _resolve_failed_entries(batch_id, entries):
    rows = marking_batch_rows(batch_id)
    by_pack = {row["pack_number"]: row["pack_number"] for row in rows}
    by_code = {code_fingerprint(row["raw_code"]): row["pack_number"] for row in rows}
    result, unknown = set(), []
    for entry in entries:
        value = str(entry or "").strip()
        if not value:
            continue
        if value in by_pack:
            result.add(value)
        else:
            match = by_code.get(code_fingerprint(value))
            if match:
                result.add(match)
            else:
                unknown.append(value)
    if unknown:
        raise RuntimeError(f"В партии не найдено: {', '.join(unknown[:5])}")
    if not result:
        raise RuntimeError("Список неуспешных кодов пуст.")
    return sorted(result)


async def batch_errors_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    document = update.message.document
    if int(document.file_size or 0) > 5 * 1024 * 1024:
        await update.message.reply_text("Файл больше 5 МБ. Пришлите короткий TXT, CSV или XLSX.")
        return BATCH_ERRORS
    telegram_file = await context.bot.get_file(document.file_id)
    content = bytes(await telegram_file.download_as_bytearray())
    try:
        if str(document.file_name or "").lower().endswith(".xlsx"):
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            entries = [str(cell) for row in workbook.active.iter_rows(values_only=True) for cell in row if cell is not None]
        else:
            text = content.decode("utf-8-sig")
            entries = [value for value in re.split(r"[,;\s]+", text) if value]
        failed = _resolve_failed_entries(context.user_data["lamoda_batch_id"], entries)
    except Exception as error:
        await update.message.reply_text(f"⚠️ Не удалось разобрать файл: {error}")
        return BATCH_ERRORS
    return await _confirm_batch_errors(update, context, failed)


async def _confirm_batch_errors(update, context, failed):
    batch_id = context.user_data["lamoda_batch_id"]
    try:
        status = confirm_marking_batch(
            batch_id, update.effective_user.id, failed,
            manager_name=_employee_name(update.effective_user),
        )
        await update.message.reply_text(f"✅ Партия №{batch_id} закрыта со статусом {status}. Ошибок: {len(failed)}.", reply_markup=back_menu())
    except Exception as error:
        await update.message.reply_text(f"⚠️ {error}")
        return BATCH_ERRORS
    return ConversationHandler.END


async def batch_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if not _is_lamoda_manager(update.effective_user):
        return
    batch_id = int(query.data.rsplit(":", 1)[1])
    try:
        cancel_marking_batch(batch_id, update.effective_user.id)
        await query.edit_message_text(f"✅ Выгрузка партии №{batch_id} отменена.", reply_markup=back_menu())
    except Exception as error:
        await _answer_error(update, error)


async def marking_history_open(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if not _is_lamoda_manager(update.effective_user):
        return
    rows = marking_history()
    lines = [f"• №{row['id']} · {row['type']} · {row['status']} · {row['created_at']:%d.%m.%Y %H:%M}" for row in rows]
    await query.edit_message_text("📚 История операций ЧЗ Lamoda\n\n" + ("\n".join(lines) if lines else "Операций нет."), reply_markup=back_menu())


def get_lamoda_handlers():
    workflow = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(assembly_continue, pattern=r"^lamoda:assembly:continue$"),
            CallbackQueryHandler(cargo_new, pattern=r"^lamoda:cargo:new:\d+$"),
            CallbackQueryHandler(cargo_reopen, pattern=r"^lamoda:cargo:reopen:\d+$"),
            CallbackQueryHandler(return_start, pattern=r"^lamoda:return:start$"),
            CallbackQueryHandler(batch_errors_start, pattern=r"^lamoda:batch:errors:\d+$"),
        ],
        states={
            SCAN_PACK: [MessageHandler(filters.TEXT & ~filters.COMMAND, scan_pack)],
            SCAN_KIZ: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, scan_kiz),
                CallbackQueryHandler(scan_kiz_skip, pattern=r"^lamoda:kiz:skip:.+$"),
            ],
            CARGO_SCAN: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, cargo_scan),
                CallbackQueryHandler(cargo_close, pattern=r"^lamoda:cargo:close:\d+$"),
            ],
            RETURN_PHOTO: [MessageHandler(filters.PHOTO, return_photo)],
            RETURN_BARCODE: [MessageHandler(filters.TEXT & ~filters.COMMAND, return_barcode)],
            RETURN_MANUAL_ORDER: [MessageHandler(filters.TEXT & ~filters.COMMAND, return_manual_order)],
            RETURN_KIZ: [MessageHandler(filters.TEXT & ~filters.COMMAND, return_kiz)],
            RETURN_CONDITION: [CallbackQueryHandler(return_condition_selected, pattern=r"^lamoda:return:condition:")],
            RETURN_DEFECT_REASON: [MessageHandler(filters.TEXT & ~filters.COMMAND, return_defect_reason)],
            RETURN_DEFECT_PHOTOS: [
                MessageHandler(filters.PHOTO, return_defect_photo),
                CallbackQueryHandler(return_defect_done, pattern=r"^lamoda:return:defect:done$"),
            ],
            BATCH_ERRORS: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, batch_errors_received),
                MessageHandler(filters.Document.ALL, batch_errors_document),
            ],
        },
        fallbacks=[], allow_reentry=True,
    )
    return [
        workflow,
        CallbackQueryHandler(assembly_start, pattern=r"^lamoda:assembly:start$"),
        CallbackQueryHandler(assembly_prepare, pattern=r"^lamoda:assembly:prepare$"),
        CallbackQueryHandler(assembly_labels_retry, pattern=r"^lamoda:labels:retry:\d+$"),
        CallbackQueryHandler(cargo_menu, pattern=r"^lamoda:cargo:menu$"),
        CallbackQueryHandler(shipment_create, pattern=r"^lamoda:shipment:create:\d+$"),
        CallbackQueryHandler(shipments_menu, pattern=r"^lamoda:shipments$"),
        CallbackQueryHandler(shipment_docs_repeat, pattern=r"^lamoda:shipment:docs:\d+$"),
        CallbackQueryHandler(returns_open, pattern=r"^lamoda:returns$"),
        CallbackQueryHandler(returns_list, pattern=r"^lamoda:return:(expected|recent|problems)$"),
        CallbackQueryHandler(marking_open, pattern=r"^lamoda:marking$"),
        CallbackQueryHandler(marking_export, pattern=r"^lamoda:marking:export:(WITHDRAWAL|REINTRODUCTION)$"),
        CallbackQueryHandler(batch_confirm, pattern=r"^lamoda:batch:confirm:\d+$"),
        CallbackQueryHandler(batch_cancel, pattern=r"^lamoda:batch:cancel:\d+$"),
        CallbackQueryHandler(marking_history_open, pattern=r"^lamoda:marking:history$"),
    ]
