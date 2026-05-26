from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from schedule_config import date_to_str, day_label, format_week_range
from schedule_google_sheets import get_schedule_matrix


GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
BLUE_FILL = PatternFill("solid", fgColor="BDD7EE")
HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")


def create_schedule_xlsx(week_start, output_path):
    output_path = Path(output_path)
    employees, dates, schedule, duty_by_date = get_schedule_matrix(week_start)

    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"

    ws.cell(row=1, column=1, value=f"Расписание {format_week_range(week_start)}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws.cell(row=2, column=1, value="ФИО сотрудника")
    for col, day in enumerate(dates, start=2):
        ws.cell(row=2, column=col, value=day_label(day))

    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, employee in enumerate(employees, start=3):
        name_cell = ws.cell(row=row_index, column=1, value=employee["full_name"])
        name_cell.font = Font(bold=True)
        name_cell.fill = WHITE_FILL

        for col, day in enumerate(dates, start=2):
            date_str = date_to_str(day)
            shift_time = schedule.get(employee["employee_id"], {}).get(date_str, "")
            is_duty = duty_by_date.get(date_str) == employee["employee_id"] and bool(shift_time)
            value = shift_time
            if is_duty:
                value = f"{shift_time}\nДежурный"

            cell = ws.cell(row=row_index, column=col, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if is_duty:
                cell.fill = BLUE_FILL
            elif shift_time:
                cell.fill = GREEN_FILL
            else:
                cell.fill = RED_FILL

    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 28
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16
    for row in range(3, 3 + len(employees)):
        ws.row_dimensions[row].height = 34

    wb.save(output_path)
    return output_path
