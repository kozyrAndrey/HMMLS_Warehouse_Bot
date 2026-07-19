import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from modules.marking.export import build_trend_island_upd_rows
from modules.marking.one_c_export import (
    BRAND_NAME,
    ONE_C_HEADERS,
    OneCExportValidationError,
    build_one_c_export_items,
    normalize_category,
    normalize_gender,
    render_one_c_xlsx,
)


GTIN = "04670332747735"
CATALOG_NAME = 'РУБАШКА "DIAMOND SHIRT BLACK" ЧЕРНАЯ L'


def make_row(**overrides):
    row = {
        "name": "DIAMOND SHIRT (L)",
        "article": "DS-L",
        "gtin": GTIN,
        "sale_price": "8490",
        "codes": ["CODE-1", "CODE-2"],
        "assortment": {
            "characteristics": [
                {"name": "Артикул", "value": "DS-L"},
                {"name": "Пол", "value": "Унисекс"},
                {"name": "Размер", "value": "L"},
                {"name": "Цвет", "value": "черный"},
                {"name": "Состав", "value": "Полиэстер 80%, хлопок 20%"},
            ],
            "barcodes": [{"ean13": "0123456789012"}, {"gtin": GTIN}],
            "salePrices": [
                {"priceType": {"name": "Цена продажи"}, "value": 849000},
            ],
        },
        "product": {
            "attributes": [
                {
                    "name": "Производитель",
                    "value": {"name": "ООО Производитель, Москва"},
                }
            ],
            "country": {"name": "RU"},
            "pathName": "hommeplusless/SHIRT",
            "salePrices": [],
        },
    }
    row.update(overrides)
    return row


class MarkingOneCExportTests(unittest.TestCase):
    def test_builds_one_row_for_one_modification(self):
        items = build_one_c_export_items([make_row()], {GTIN: CATALOG_NAME})

        self.assertEqual(len(items), 1)
        item = items[0]
        self.assertEqual(item.article, "DS-L")
        self.assertEqual(item.honest_sign_name, CATALOG_NAME)
        self.assertNotIn("DS-L", item.honest_sign_name)
        self.assertEqual(item.brand, BRAND_NAME)
        self.assertEqual(item.marking_code_count, 2)
        self.assertEqual(item.ean13, "0123456789012")
        self.assertEqual(item.category, "Товары легкой промышленности")

    def test_normalizes_all_gender_values(self):
        cases = {
            "муж.": "Мужской",
            "Для мужчин": "Мужской",
            "ЖЕНСКИЙ": "Женский",
            "для женщин": "Женский",
            "унисекс": "Unisex",
            "UNISEX": "Unisex",
            "дети": "Детский",
            "для детей": "Детский",
            "Не применимо": "Не применимо",
        }
        for source, expected in cases.items():
            with self.subTest(source=source):
                self.assertEqual(normalize_gender(source), expected)

    def test_unknown_gender_is_error(self):
        with self.assertRaises(ValueError):
            normalize_gender("Для взрослых")

    def test_normalizes_all_categories(self):
        cases = {
            "Одежда / футболки": "Товары легкой промышленности",
            "SHOES / Кроссовки": "Обувь",
            "Парфюмерия и духи": "Духи и туалетная вода",
            "Изделия из натурального меха": "Изделия из меха",
        }
        for source, expected in cases.items():
            with self.subTest(source=source):
                self.assertEqual(normalize_category(source), expected)

    def test_unknown_category_is_error(self):
        with self.assertRaises(ValueError):
            normalize_category("Электроника")

    def test_export_always_uses_fixed_light_industry_category(self):
        row = make_row()
        row["product"]["attributes"].append(
            {"name": "Категория изделия", "value": "Толстовка"}
        )
        row["product"]["pathName"] = "Нераспознанная группа"

        item = build_one_c_export_items([row], {GTIN: CATALOG_NAME})[0]

        self.assertEqual(item.category, "Товары легкой промышленности")

    def test_variant_characteristics_override_product_attributes(self):
        row = make_row()
        row["product"]["attributes"].extend(
            [
                {"name": "Размер", "value": "XL"},
                {"name": "Цвет", "value": "белый"},
                {"name": "Состав", "value": "Хлопок 100%"},
            ]
        )

        item = build_one_c_export_items([row], {GTIN: CATALOG_NAME})[0]

        self.assertEqual(item.size, "L")
        self.assertEqual(item.color, "черный")
        self.assertEqual(item.composition, "Полиэстер 80%, хлопок 20%")

    def test_counts_only_unique_marking_codes_for_same_modification(self):
        first = make_row(codes=["CODE-1", "CODE-1", "CODE-2"])
        second = make_row(codes=["CODE-2", "CODE-3"])

        item = build_one_c_export_items([first, second], {GTIN: CATALOG_NAME})[0]

        self.assertEqual(item.marking_code_count, 3)

    def test_preserves_leading_zero_in_ean13(self):
        item = build_one_c_export_items([make_row()], {GTIN: CATALOG_NAME})[0]
        self.assertEqual(item.ean13, "0123456789012")

    def test_prefers_exact_retail_price(self):
        row = make_row()
        row["assortment"]["salePrices"] = [
            {"priceType": {"name": "Цена продажи"}, "value": 849000},
            {"priceType": {"name": "Розничная цена"}, "value": 899050},
        ]

        item = build_one_c_export_items([row], {GTIN: CATALOG_NAME})[0]

        self.assertEqual(str(item.retail_price), "8990.50")

    def test_multiple_ean13_values_are_error(self):
        row = make_row()
        row["assortment"]["barcodes"].append({"ean13": "0123456789029"})

        with self.assertRaises(OneCExportValidationError) as caught:
            build_one_c_export_items([row], {GTIN: CATALOG_NAME})

        self.assertTrue(any("несколько разных EAN-13" in error for error in caught.exception.errors))

    def test_missing_required_field_is_error(self):
        row = make_row()
        row["product"]["attributes"] = []

        with self.assertRaises(OneCExportValidationError) as caught:
            build_one_c_export_items([row], {GTIN: CATALOG_NAME})

        self.assertTrue(any("отсутствует производитель" in error for error in caught.exception.errors))

    def test_csv_can_succeed_when_one_c_validation_fails(self):
        row = make_row()
        row["product"]["attributes"] = []

        csv_rows = build_trend_island_upd_rows([row], {GTIN: CATALOG_NAME})
        with self.assertRaises(OneCExportValidationError):
            build_one_c_export_items([row], {GTIN: CATALOG_NAME})

        self.assertEqual(len(csv_rows), 2)

    def test_different_sizes_are_not_merged(self):
        second = make_row(
            article="DS-M",
            gtin="04670332747728",
            codes=["CODE-3"],
            assortment={
                **make_row()["assortment"],
                "characteristics": [
                    {"name": "Артикул", "value": "DS-M"},
                    {"name": "Пол", "value": "Унисекс"},
                    {"name": "Размер", "value": "M"},
                    {"name": "Цвет", "value": "черный"},
                    {"name": "Состав", "value": "Полиэстер 80%, хлопок 20%"},
                ],
                "barcodes": [{"ean13": "0123456789029"}, {"gtin": "04670332747728"}],
            },
        )

        items = build_one_c_export_items(
            [make_row(), second],
            {GTIN: CATALOG_NAME, "04670332747728": "РУБАШКА ЧЕРНАЯ M"},
        )

        self.assertEqual(len(items), 2)
        self.assertEqual({item.size for item in items}, {"L", "M"})

    def test_same_gtin_with_different_ean13_is_error(self):
        second = make_row(codes=["CODE-3"])
        second["assortment"] = {
            **second["assortment"],
            "barcodes": [{"ean13": "0123456789029"}, {"gtin": GTIN}],
        }

        with self.assertRaises(OneCExportValidationError) as caught:
            build_one_c_export_items([make_row(), second], {GTIN: CATALOG_NAME})

        self.assertTrue(any("одному GTIN" in error for error in caught.exception.errors))

    def test_same_marking_code_for_different_modifications_is_error(self):
        second = make_row(article="DS-M", gtin="04670332747728", codes=["CODE-1"])
        second["assortment"] = {
            **second["assortment"],
            "characteristics": [
                {"name": "Артикул", "value": "DS-M"},
                {"name": "Пол", "value": "Унисекс"},
                {"name": "Размер", "value": "M"},
                {"name": "Цвет", "value": "черный"},
                {"name": "Состав", "value": "Полиэстер 80%, хлопок 20%"},
            ],
            "barcodes": [{"ean13": "0123456789029"}, {"gtin": "04670332747728"}],
        }

        with self.assertRaises(OneCExportValidationError) as caught:
            build_one_c_export_items(
                [make_row(), second],
                {GTIN: CATALOG_NAME, "04670332747728": "РУБАШКА ЧЕРНАЯ M"},
            )

        self.assertTrue(any("к нескольким товарным модификациям" in error for error in caught.exception.errors))

    def test_creates_real_xlsx_matching_template(self):
        item = build_one_c_export_items([make_row()], {GTIN: CATALOG_NAME})[0]
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "export.xlsx"
            render_one_c_xlsx([item], path)
            self.assertTrue(zipfile.is_zipfile(path))
            workbook = load_workbook(path, data_only=False)
            worksheet = workbook["Список товаров"]

        self.assertEqual(workbook.sheetnames, ["Список товаров", "Категория", "Классификатор стран", "Пол список"])
        self.assertEqual([worksheet.cell(3, column).value for column in range(1, 15)], ONE_C_HEADERS)
        self.assertEqual(worksheet.max_row, 4)
        self.assertEqual(worksheet["B4"].value, "DS-L")
        self.assertEqual(worksheet["C4"].value, CATALOG_NAME)
        self.assertEqual(worksheet["F4"].value, "Hommeplusless")
        self.assertEqual(worksheet["M4"].value, "0123456789012")
        self.assertEqual(worksheet["M4"].number_format, "@")
        self.assertEqual(worksheet["K4"].data_type, "n")
        self.assertTrue(all(sheet.sheet_state == "hidden" for sheet in workbook.worksheets[1:]))


if __name__ == "__main__":
    unittest.main()
