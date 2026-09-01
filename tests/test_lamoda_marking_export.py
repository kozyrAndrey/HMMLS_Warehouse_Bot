import unittest
from io import BytesIO

from openpyxl import load_workbook

from modules.lamoda_fbs.pdf_reports import create_marking_xlsx
from modules.lamoda_fbs.storage import sold_price_from_item_json


class LamodaMarkingExportTests(unittest.TestCase):
    def test_paid_price_is_preferred_and_converted_from_minor_units(self):
        raw_json = (
            '{"paidPrice":{"amount":2599000,"currency":"RUB"},'
            '"salePrice":{"amount":2999000,"currency":"RUB"}}'
        )

        self.assertEqual(sold_price_from_item_json(raw_json), 25990)

    def test_withdrawal_xlsx_contains_only_requested_columns(self):
        content = create_marking_xlsx(
            [{
                "item_id": "ITEM-001",
                "product_name": "Футболка",
                "short_code": "010460123456789321ABCDEFGHIJKLM",
                "sale_price": 25990,
                "order_id": "ORDER-SHOULD-NOT-BE-EXPORTED",
                "raw_code": "FULL-CODE-SHOULD-NOT-BE-EXPORTED",
            }],
            "WITHDRAWAL",
        )

        sheet = load_workbook(BytesIO(content)).active
        self.assertEqual(
            [cell.value for cell in sheet[1]],
            ["Номер товарной этикетки", "Название товара", "Код маркировки", "Цена продажи"],
        )
        self.assertEqual(
            [cell.value for cell in sheet[2]],
            ["ITEM-001", "Футболка", "010460123456789321ABCDEFGHIJKLM", 25990],
        )
        self.assertEqual(sheet.max_column, 4)


if __name__ == "__main__":
    unittest.main()
