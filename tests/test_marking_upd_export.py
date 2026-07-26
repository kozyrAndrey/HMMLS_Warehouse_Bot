import csv
import tempfile
import unittest
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from modules.marking.export import (
    TrendExportValidationError,
    build_trend_island_upd_rows,
    calculate_net_price,
    create_stock_marking_codes_xlsx,
    create_trend_island_upd_csv,
    extract_article,
    extract_gtin,
    extract_sale_price,
    get_unmarked_moysklad_rows,
)
from modules.marking.storage import normalize_gtin


class MarkingUpdExportTests(unittest.TestCase):
    def test_stock_codes_xlsx_contains_only_codes_without_header(self):
        rows = [
            {"codes": ["CODE-1", "CODE-2"]},
            {"codes": ["CODE-3"]},
        ]

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "stock_codes.xlsx"
            create_stock_marking_codes_xlsx(rows, path)
            worksheet = load_workbook(path)["Коды маркировки"]

        self.assertEqual(worksheet.max_column, 1)
        self.assertEqual(
            [worksheet.cell(row=index, column=1).value for index in range(1, 4)],
            ["CODE-1", "CODE-2", "CODE-3"],
        )

    def test_price_without_vat_has_eight_decimal_places(self):
        self.assertEqual(calculate_net_price("49.00"), Decimal("45.79439252"))
        self.assertEqual(calculate_net_price("11890"), Decimal("11112.14953271"))

    def test_builds_one_row_per_code_and_matches_gtin13_to_gtin14(self):
        code_one = "010467033274423921ABC123XYZ\x1d91AA"
        code_two = "010467033274423921ABC456XYZ"
        rows = [
            {
                "name": "DIAMOND HOODIE BLACK S",
                "article": "ART-1",
                "gtin": "4670332744239",
                "sale_price": Decimal("11890"),
                "codes": [code_one, code_two],
            }
        ]

        result = build_trend_island_upd_rows(
            rows,
            {"04670332744239": 'ХУДИ "DIAMOND HOODIE BLACK", ЧЕРНЫЙ S'},
        )

        self.assertEqual(len(result), 2)
        self.assertEqual(
            result[0],
            [
                "1",
                'ART-1 ХУДИ "DIAMOND HOODIE BLACK", ЧЕРНЫЙ S',
                "11112.14953271",
                "2",
                "796",
                "7%",
                "КИЗ",
                code_one,
            ],
        )
        self.assertEqual(result[1][0], "1")
        self.assertEqual(result[1][7], code_two)

    def test_numbers_all_codes_of_the_same_gtin_identically(self):
        rows = [
            {
                "name": "Другой товар",
                "article": "ART-2",
                "gtin": "4670332747734",
                "sale_price": "49",
                "codes": ["CODE-3"],
            },
            {
                "name": "Тот же товар",
                "article": "ART-1",
                "gtin": "04670332744239",
                "sale_price": "49",
                "codes": ["CODE-2"],
            },
            {
                "name": "Первый товар",
                "article": "ART-1",
                "gtin": "4670332744239",
                "sale_price": "49",
                "codes": ["CODE-1"],
            },
        ]

        result = build_trend_island_upd_rows(
            rows,
            {
                "4670332744239": "Первый товар",
                "4670332747734": "Другой товар",
            },
        )

        self.assertEqual([row[0] for row in result], ["1", "1", "2"])
        self.assertEqual([row[3] for row in result], ["2", "2", "1"])
        self.assertEqual([row[1] for row in result], [
            "ART-1 Первый товар",
            "ART-1 Первый товар",
            "ART-2 Другой товар",
        ])

    def test_csv_matches_reference_dialect_and_preserves_code(self):
        code = "010467033274423921ABC123XYZ"
        rows = [
            {
                "name": "Товар",
                "article": "ART-1",
                "gtin": "4670332744239",
                "sale_price": "49",
                "codes": [code],
            }
        ]

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "export.csv"
            create_trend_island_upd_csv(rows, {"4670332744239": 'Название, "модель"'}, path)
            raw = path.read_bytes()
            self.assertFalse(raw.startswith(b"\xef\xbb\xbf"))
            self.assertNotIn(b"\r\n", raw)
            with path.open(encoding="utf-8", newline="") as file:
                parsed = list(csv.reader(file))

        self.assertEqual(len(parsed), 1)
        self.assertEqual(len(parsed[0]), 8)
        self.assertEqual(parsed[0][1], 'ART-1 Название, "модель"')
        self.assertEqual(parsed[0][7], code)

    def test_reports_all_required_field_errors(self):
        rows = [
            {
                "name": "Проблемный товар",
                "article": "",
                "gtin": "",
                "sale_price": None,
                "codes": ["CODE-1"],
            }
        ]

        with self.assertRaises(TrendExportValidationError) as caught:
            build_trend_island_upd_rows(rows, {})

        message = "\n".join(caught.exception.errors)
        self.assertIn("отсутствует артикул", message)
        self.assertIn("отсутствует GTIN", message)
        self.assertIn("отсутствует цена продажи", message)

    def test_skips_unmarked_position_in_upd_csv(self):
        rows = [
            {
                "name": "HOMME BIRKIN MESSENGER BAG",
                "article": "BAG-1",
                "gtin": "2000000000001",
                "sale_price": "14990",
                "codes": [],
            }
        ]

        self.assertEqual(build_trend_island_upd_rows(rows, {}), [])

    def test_marked_catalog_position_without_codes_is_error(self):
        rows = [
            {
                "name": "Маркируемый товар",
                "article": "ART-1",
                "gtin": "4670332744239",
                "sale_price": "49",
                "codes": [],
            }
        ]

        with self.assertRaises(TrendExportValidationError) as caught:
            build_trend_island_upd_rows(
                rows,
                {"4670332744239": "Маркируемый товар"},
            )

        self.assertTrue(
            any("отсутствуют коды маркировки" in error for error in caught.exception.errors)
        )

    def test_reports_missing_catalog_mapping(self):
        rows = [
            {
                "name": "Товар без сопоставления",
                "article": "ART-1",
                "gtin": "4670332744239",
                "sale_price": "49",
                "codes": ["CODE-1"],
            }
        ]

        with self.assertRaises(TrendExportValidationError) as caught:
            build_trend_island_upd_rows(rows, {})

        self.assertIn("отсутствует в справочнике", caught.exception.errors[0])

    def test_reports_duplicate_marking_code(self):
        rows = [
            {
                "name": "Товар 1",
                "article": "A-1",
                "gtin": "4670332744239",
                "sale_price": "49",
                "codes": ["DUPLICATE"],
            },
            {
                "name": "Товар 2",
                "article": "A-2",
                "gtin": "4670332744239",
                "sale_price": "49",
                "codes": ["DUPLICATE"],
            },
        ]

        with self.assertRaises(TrendExportValidationError) as caught:
            build_trend_island_upd_rows(rows, {"4670332744239": "Название"})

        self.assertTrue(any("встречается несколько раз" in error for error in caught.exception.errors))

    def test_extracts_variant_gtin_and_parent_article_and_sale_price(self):
        assortment = {
            "barcodes": [{"ean13": "4670332744239"}],
            "salePrices": [
                {"priceType": {"name": "Цена продажи"}, "value": 1189000},
                {"priceType": {"name": "Старая цена"}, "value": 1299000},
            ],
        }
        parent = {"article": "ART-1", "barcodes": [{"gtin": "04670332744239"}]}

        self.assertEqual(extract_article(assortment, parent), "ART-1")
        self.assertEqual(extract_gtin(assortment, parent), "4670332744239")
        self.assertEqual(extract_sale_price(assortment, parent, "Цена продажи"), Decimal("11890"))
        self.assertEqual(extract_sale_price(assortment, parent, "Старая цена"), Decimal("12990"))

    def test_prefers_size_specific_article_from_variant_characteristics(self):
        assortment = {
            "article": None,
            "characteristics": [
                {"name": "Размер", "value": "L"},
                {"name": "Артикул", "value": "DS-L"},
            ],
        }
        parent = {"article": "DS"}

        self.assertEqual(extract_article(assortment, parent), "DS-L")

    def test_normalize_gtin_uses_gtin14_for_matching(self):
        self.assertEqual(normalize_gtin("4670332744239"), "04670332744239")
        self.assertEqual(normalize_gtin("04670332744239"), "04670332744239")

    def test_finds_unmarked_product_in_moysklad_assortment_by_model_name(self):
        assortment = {
            "name": "HOMME BIRKIN MESSENGER BAG",
            "article": "BAG-1",
            "barcodes": [{"ean13": "2000000000001"}],
            "salePrices": [
                {"priceType": {"name": "Цена продажи"}, "value": 1499000}
            ],
        }

        class Client:
            def list_entities(self, entity_type, params=None):
                self.entity_type = entity_type
                self.params = params
                return {"rows": [assortment]}

        client = Client()
        rows = get_unmarked_moysklad_rows(
            client,
            [
                {
                    "honest_sign_name": (
                        'СУМКА МЕССЕНДЖЕР "HOMME BIRKIN MESSENGER" ЧЕРНАЯ'
                    )
                }
            ],
            sale_price_type="Цена продажи",
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["name"], "HOMME BIRKIN MESSENGER BAG")
        self.assertEqual(rows[0]["article"], "BAG-1")
        self.assertEqual(rows[0]["sale_price"], Decimal("14990"))
        self.assertEqual(rows[0]["codes"], [])
        self.assertEqual(client.entity_type, "assortment")
        self.assertEqual(client.params["limit"], 1000)
        self.assertEqual(client.params["offset"], 0)


if __name__ == "__main__":
    unittest.main()
